from openpyxl import *
from openpyxl.utils import *
from random import *

# 엑셀 파일을 불러옵니다.
wb = load_workbook('example.xlsx')
# 엑셀에 새로운 시트를 추가하고, 시트를 가져옵니다.
ws = wb.get_sheet_by_name("NGM")
# 추가한 시트의 탭 컬러를 Yellow로 설정합니다.
# ws.sheet_properties.tabColor = 'ffff00'

# for r in range(1, 11):
#     for c in range(1, 11):
#         ws.cell(row=r, column=c, value=randint(50, 100))

# 합계를 계산하는 수식을 넣어줍니다.
for c in range(1, 11):
    # 합계를 표시할 셀을 가져옵니다.
    cell = ws.cell(row=11, column=c)
    print(cell)
    ws.cell(row=11, column=c, value='SUM')
    # 컬럼의 레터를 알아내기 위한 함수입니다.
    letter = get_column_letter(c)
    # 합계를 12번째 로우에 추가 해줍니다.
    ws.cell(row=12, column=c, value='=sum('+letter+'1:'+letter+'11)')
    ws.cell(13, c, 'EXE')

# 워크북을 저장합니다. (Excel 저장)
wb.save('example.xlsx')
# 열려 있으면 엑셀 파일을 닫습니다. 
# 읽기 전용 및 쓰기 전용 모드에만 영향을 줍니다.
wb.close()